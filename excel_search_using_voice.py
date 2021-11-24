import speech_recognition as sr
import openpyxl

r = sr.Recognizer()
wb = openpyxl.load_workbook("Hello.xlsx")
ws = wb.active

with sr.Microphone() as source:
    print("Talk")
    audio_text = r.listen(source)
    print("Time over, thanks")
# recoginize_() method will throw a request error if the API is unreachable, hence using exception handling
    
    try:
        # using google speech recognition
        print("Text: "+r.recognize_google(audio_text))
    except:
        print("Sorry, I did not get that")

word=r.recognize_google(audio_text)

for r in range(1,ws.max_row+1):
    for c in range(1,ws.max_column+1):
        s = ws.cell(r,c).value
        cell_v=ws.cell(row= r,column=7)
        if str(s)==word:
           cell_v.value="YES"
           print("updated")
cell_Y=ws.cell(row= 1,column=7)  
cell_Y.value="PRESENT"      
wb.save('Hello.xlsx')

