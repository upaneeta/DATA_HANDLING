import openpyxl

path = "HELLO.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# printing any particular cell value
r=int(input("Enter the row number:"))
c=int(input("Enter the coloumn number:"))
cell_obj = sheet_obj.cell(row = r, column = c)
print(cell_obj.value)  # printing any particular cell value 

# maximum no. of rows
mxr=(sheet_obj.max_row) # maximum no. of rows
print(mxr)

# maximum no. of columns
mxc=sheet_obj.max_column # maximum no. of columns
print(mxc)

# print any row
for i in range(1, mxc+1):
    row_all=sheet_obj.cell(row=r,column=i)
    print(row_all.value, end=" ")
    
# print any column
for j in range(1, mxr+1):
    col_all=sheet_obj.cell(row=j,column=c)
    print("\n"+col_all.value)
